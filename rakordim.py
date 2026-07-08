"""
Rakordim mujor: Libri i Shitjeve Financa vs Portali Self Care
Identifikon automatikisht:
 1. Diferencat në faturat me subjekt të identifikuar (match sipas Numrit Serial/NIVF)
 2. Diferencat në totalet ditore të klientëve të rastësishëm
 3. Operatorin (pikën e shitjes) përgjegjës për diferencën, në ditët problematike
"""
import pandas as pd
import numpy as np
from itertools import combinations
import sys

TOLERANCE = 5  # lek, diferenca nën këtë vlerë konsiderohen rrumbullakosje

# Kolonat e kategorive te TVSH-se ne Librin e Shitjes (format zyrtar), sipas kodeve
# a,b,c,ç,d,dh,e,ë,f,g,gj,h,i,j,k,l,ll,m,n,nj,o,p,q -> indekset 0-22.
# 'e' (indeks 6, Vlera totale) tashme perfshihet si 'Vlera'; kategorite e meposhtme
# jane komponentet qe e perbejne ate total - nese nje shitje kalon gabimisht nga
# nje kategori ne tjetren me te njejten vlere, totali nuk e kap gabimin, prandaj
# krahasohen edhe keto ndaras.
KATEGORITE_TVSH = {
    7: 'Perjashtuar_e',
    8: 'PaTvsh_f',
    9: 'PaTvsh_g',
    10: 'PaTvsh_gj',
    11: 'Tatueshme20',
    12: 'Tvsh20',
    13: 'Tatueshme10',
    14: 'Tvsh10',
    15: 'Tatueshme6',
    16: 'Tvsh6',
    17: 'Import20_baze',
    18: 'Import20_tvsh',
    19: 'ImportTjeter_baze',
    20: 'ImportTjeter_tvsh',
    21: 'Tjeter_p',
    22: 'Tjeter_q',
}

def parse_libri(path, sheet_name):
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None)
    code_row_idx = raw[raw[0].astype(str) == 'a'].index[0]
    stop_candidates = raw[raw[0].astype(str).str.contains('Shuma', case=False, na=False)].index
    stop_idx = stop_candidates[stop_candidates > code_row_idx][0]
    kat_cols = list(KATEGORITE_TVSH.keys())
    data = raw.iloc[code_row_idx+1:stop_idx, [0,1,2,3,4,5,6]+kat_cols].copy()
    data.columns = ['NrFature','NumriSerial','Data','Bleresi','Rrethi','NIPT','Vlera'] + list(KATEGORITE_TVSH.values())
    data['Data'] = pd.to_datetime(data['Data'], errors='coerce', dayfirst=True)
    data['Vlera'] = pd.to_numeric(data['Vlera'], errors='coerce')
    for col in KATEGORITE_TVSH.values():
        data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0)
    data['NIPT'] = data['NIPT'].astype(str).str.strip()
    data.loc[data['NIPT'].isin(['nan','None','']), 'NIPT'] = np.nan
    return data.reset_index(drop=True)


def krahaso_kategorite_tvsh(fin, port, tol=TOLERANCE):
    """
    Krahason totalet ditore te cdo kategorie TVSH-je (jo vetem totalin final).
    Kap rastet kur nje shitje kalon gabimisht nga nje kategori (p.sh. e perjashtuar
    nga TVSH) ne nje tjeter (p.sh. e tatueshme 20%) me te njejten vlere - gabim qe
    krahasimi i totalit te vetem nuk e sheh.
    """
    kategori_cols = list(KATEGORITE_TVSH.values())
    fin_by_date = fin.groupby('Data')[kategori_cols].sum()
    port_by_date = port.groupby('Data')[kategori_cols].sum()

    rows = []
    for data_val in sorted(set(fin_by_date.index) | set(port_by_date.index)):
        fin_row = fin_by_date.loc[data_val] if data_val in fin_by_date.index else pd.Series(0, index=kategori_cols)
        port_row = port_by_date.loc[data_val] if data_val in port_by_date.index else pd.Series(0, index=kategori_cols)
        for kat in kategori_cols:
            diff = fin_row[kat] - port_row[kat]
            if abs(diff) > tol:
                rows.append({
                    'Data': data_val, 'Kategoria': kat,
                    'Financa': fin_row[kat], 'Portal': port_row[kat], 'Diferenca': diff
                })
    return pd.DataFrame(rows, columns=['Data','Kategoria','Financa','Portal','Diferenca'])


def parse_transaksionet(path):
    """Lexon skedarin e transaksioneve te detajuara te eksportuar nga Portali Self Care (fleta 'data')."""
    df = pd.read_excel(path, sheet_name='data')
    df['Data e fiskalizimit'] = pd.to_datetime(df['Data e fiskalizimit'], errors='coerce', dayfirst=True)
    return df


def identifiko_operatorin_transaksione(fin_rand, trans_df, tol=5):
    """
    Perputh cdo pike shitjeje (Bleresi ne Financa) me kodin e operatorit ne Portal,
    duke krahasuar totalet ditore te klienteve te rastesishem (pa NIVF).
    Kthen nje DataFrame me diferencat per cdo pike shitjeje te dites.
    """
    data_val = trans_df['Data e fiskalizimit'].mode().iloc[0]
    random_tx = trans_df[trans_df['Numri i identifikimit të blerësit'].isna()]
    portal_by_code = random_tx.groupby('Kodi i operatorit')['Vlera totale me TVSH'].sum().to_dict()

    fin_day = fin_rand[fin_rand['Data'] == data_val]
    fin_by_bleresi = fin_day.groupby('Bleresi')['Vlera'].sum().to_dict()

    pairs = [
        (abs(fv - pv), bleresi, code)
        for bleresi, fv in fin_by_bleresi.items()
        for code, pv in portal_by_code.items()
    ]
    pairs.sort(key=lambda x: x[0])

    matched_bleresi, matched_code = {}, {}
    for _, bleresi, code in pairs:
        if bleresi in matched_bleresi or code in matched_code:
            continue
        matched_bleresi[bleresi] = code
        matched_code[code] = bleresi

    rows = []
    for bleresi, fin_val in fin_by_bleresi.items():
        code = matched_bleresi.get(bleresi)
        portal_val = portal_by_code.get(code) if code else None
        diff = fin_val - (portal_val if portal_val is not None else 0)
        rows.append({
            'Data': data_val, 'Bleresi': bleresi, 'Financa_Vlera': fin_val,
            'Kodi_Operatorit': code, 'Portal_Vlera': portal_val, 'Diferenca': diff
        })
    for code, portal_val in portal_by_code.items():
        if code not in matched_code:
            rows.append({
                'Data': data_val, 'Bleresi': None, 'Financa_Vlera': None,
                'Kodi_Operatorit': code, 'Portal_Vlera': portal_val, 'Diferenca': -portal_val
            })

    result = pd.DataFrame(rows)
    result['Status'] = np.where(result['Diferenca'].abs() <= tol, 'OK', 'KONTROLLO')
    return result.sort_values('Diferenca', key=abs, ascending=False).reset_index(drop=True)


def find_operator_combo(target_diff, operator_totals, tol=5, max_combo=3):
    """Gjen kombinimin e operatorëve (deri në max_combo) që shpjegon diferencën."""
    items = list(operator_totals.items())
    for r in range(1, max_combo+1):
        for combo in combinations(items, r):
            s = sum(v for _, v in combo)
            if abs(s - target_diff) <= tol:
                return [name for name, _ in combo], s
    return None, None


def rakordo(financa_path, financa_sheet, portal_path, portal_sheet):
    fin = parse_libri(financa_path, financa_sheet)
    port = parse_libri(portal_path, portal_sheet)

    # --- 1. Subjekt i identifikuar ---
    fin_id = fin[fin['NIPT'].notna()]
    port_id = port[port['NIPT'].notna()]
    merged = pd.merge(
        fin_id[['NumriSerial','Data','Bleresi','NIPT','Vlera']],
        port_id[['NumriSerial','Vlera']],
        on='NumriSerial', how='outer', suffixes=('_Financa','_Portal'), indicator=True
    )
    merged['Diferenca'] = merged['Vlera_Financa'].fillna(0) - merged['Vlera_Portal'].fillna(0)
    subjekt_diff = merged[(merged['Diferenca'].abs() > TOLERANCE) | (merged['_merge'] != 'both')].copy()
    subjekt_diff['Problemi'] = np.where(
        subjekt_diff['_merge'] == 'left_only', 'Mungon në Portal',
        np.where(subjekt_diff['_merge'] == 'right_only', 'Mungon në Financa', 'Vlerë e ndryshme')
    )

    # --- 2. Klient i rastësishëm - totale ditore ---
    fin_rand = fin[fin['NIPT'].isna()]
    port_rand = port[port['NIPT'].isna()]
    fin_by_date = fin_rand.groupby('Data')['Vlera'].sum().rename('Financa_Total')
    port_by_date = port_rand.groupby('Data')['Vlera'].sum().rename('Portal_Total')
    cmp = pd.concat([fin_by_date, port_by_date], axis=1).fillna(0)
    cmp['Diferenca'] = cmp['Financa_Total'] - cmp['Portal_Total']
    cmp['Status'] = np.where(cmp['Diferenca'].abs() <= TOLERANCE, 'OK', 'KONTROLLO')
    cmp = cmp.reset_index()

    # --- 3. Identifikimi i operatorit për datat problematike ---
    problem_rows = []
    for _, row in cmp[cmp['Status']=='KONTROLLO'].iterrows():
        data_val = row['Data']
        diff = row['Diferenca']
        day_ops = fin_rand[fin_rand['Data']==data_val].groupby('Bleresi')['Vlera'].sum().to_dict()
        combo, matched_sum = find_operator_combo(diff, day_ops, tol=TOLERANCE)
        problem_rows.append({
            'Data': data_val,
            'Diferenca': diff,
            'Operatori(et) i identifikuar': ', '.join(combo) if combo else 'I PAIDENTIFIKUAR (kontroll manual)',
            'Shuma e përputhur': matched_sum if matched_sum is not None else None
        })
    operator_report = pd.DataFrame(problem_rows)

    # --- 4. Diferenca sipas kategorive te TVSH-se (edhe kur totali ditor perputhet) ---
    kategori_diff = krahaso_kategorite_tvsh(fin, port)

    return subjekt_diff, cmp, operator_report, fin_rand, kategori_diff


def format_workbook(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    header_fill = PatternFill('solid', start_color='1F4E78', end_color='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    red_fill = PatternFill('solid', start_color='FFC7CE', end_color='FFC7CE')
    green_fill = PatternFill('solid', start_color='C6EFCE', end_color='C6EFCE')
    base_font = Font(name='Arial')

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        headers = [c.value for c in ws[1]]
        status_col = headers.index('Status')+1 if 'Status' in headers else None
        oper_col = headers.index('Operatori(et) i identifikuar')+1 if 'Operatori(et) i identifikuar' in headers else None
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = base_font
            if status_col and row[status_col-1].value == 'KONTROLLO':
                for cell in row:
                    cell.fill = red_fill
            if status_col and row[status_col-1].value == 'OK':
                for cell in row:
                    cell.fill = green_fill
            if oper_col and row[oper_col-1].value and 'PAIDENTIFIKUAR' in str(row[oper_col-1].value):
                for cell in row:
                    cell.fill = red_fill
            elif oper_col and row[oper_col-1].value:
                for cell in row:
                    cell.fill = green_fill
        for i, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len+2, 45)
        ws.freeze_panes = 'A2'
    wb.save(path)


if __name__ == '__main__':
    financa_path, portal_path, out_path = sys.argv[1], sys.argv[2], sys.argv[3]
    subjekt_diff, cmp, operator_report, fin_rand, kategori_diff = rakordo(financa_path, 'shitje', portal_path, 'Sales book')

    subjekt_diff = subjekt_diff.drop(columns=['_merge'])

    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        operator_report.to_excel(writer, sheet_name='Operatori_Problematik', index=False)
        cmp.to_excel(writer, sheet_name='Klient_Rastesishem_Ditor', index=False)
        subjekt_diff.to_excel(writer, sheet_name='Subjekt_Identifikuar', index=False)
        kategori_diff.to_excel(writer, sheet_name='Diferenca_Kategori_TVSH', index=False)

    format_workbook(out_path)

    print("Subjekt identifikuar - diferenca gjetur:", len(subjekt_diff))
    print("Ditë me diferencë tek klienti i rastësishëm:", (cmp['Status']=='KONTROLLO').sum())
    print("Diferenca sipas kategorive TVSH:", len(kategori_diff))
    print(operator_report.to_string())
    if len(kategori_diff):
        print(kategori_diff.to_string())
